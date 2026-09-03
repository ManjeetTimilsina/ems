import os
from io import BytesIO

from openpyxl import load_workbook

import app as app_module
from ems.database import get_connection, initialize_database
from ems import employees as emp
from ems import shifts as sh
from ems import schedule as sc


def setup_test_db(tmp_path, monkeypatch):
    db_path = tmp_path / 'test_ems.db'
    import ems.database as db_mod
    monkeypatch.setattr(db_mod, 'DB_PATH', str(db_path))
    initialize_database()
    return db_path


def test_dashboard_shift_panel_is_collapsible_and_searchable():
    client = app_module.app.test_client()
    response = client.get('/')
    assert response.status_code == 200
    html = response.get_data(as_text=True)

    assert 'On Shift Today' in html
    assert 'data-table-search' in html
    assert 'Department' in html
    assert '<details' in html


def test_remove_employee_pattern_schedule(tmp_path, monkeypatch):
    setup_test_db(tmp_path, monkeypatch)

    ok, msg = emp.add_employee('E-200', 'Clear', 'Schedule', department='Ops')
    assert ok, msg
    emp_id = emp.list_employees()[0]['id']
    sh.add_shift_type('Test Shift', '08:00', '16:00')
    shift_id = sh.list_shift_types()[0]['id']

    ok, msg = app_module.rec.add_pattern(
        employee_id=emp_id,
        shift_type_id=shift_id,
        weekday=0,
        effective_start='2026-09-01',
        effective_end='2026-12-31',
        frequency='weekly',
    )
    assert ok, msg

    created = app_module.rec.generate_schedule_for_month(2026, 9)
    assert created > 0

    client = app_module.app.test_client()
    response = client.post('/patterns/remove-employee', data={
        'employee_id': str(emp_id),
    }, follow_redirects=True)

    assert response.status_code == 200
    assert app_module.rec.list_patterns(employee_id=emp_id) == []
    assert len(sc.get_schedule_for_date('2026-09-05')) == 0
    assert len(sc.get_schedule_for_date('2026-09-12')) == 0
    assert len(sc.get_schedule_for_date('2026-10-10')) == 0


def test_reports_export_matches_employee_filter(tmp_path, monkeypatch):
    setup_test_db(tmp_path, monkeypatch)

    ok, msg = emp.add_employee('E-100', 'Alpha', 'One', department='Ops')
    assert ok, msg
    ok, msg = emp.add_employee('E-200', 'Bravo', 'Two', department='Ops')
    assert ok, msg
    emp1_id = emp.list_employees(status='Active')[0]['id']
    emp2_id = emp.list_employees(status='Active')[1]['id']
    sh.add_shift_type('Shift A', '08:00', '16:00')
    shift_id = sh.list_shift_types()[0]['id']

    ok, msg = sc.assign_shift(emp1_id, shift_id, '2026-09-05', status='Worked', notes='')
    assert ok, msg
    ok, msg = sc.assign_shift(emp2_id, shift_id, '2026-09-06', status='Worked', notes='')
    assert ok, msg

    client = app_module.app.test_client()
    response = client.get(f'/reports/export?year=2026&month=9&employee_id={emp1_id}')

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.data))
    rows = list(workbook.active.iter_rows(values_only=True))
    assert any(row and row[0] == 'E-100' for row in rows)
    assert not any(row and row[0] == 'E-200' for row in rows)


def test_reports_export_includes_summary_values(tmp_path, monkeypatch):
    setup_test_db(tmp_path, monkeypatch)

    ok, msg = emp.add_employee('E-300', 'Gamma', 'Three', department='Ops', position='Nurse')
    assert ok, msg
    emp_id = emp.list_employees(status='Active')[0]['id']
    sh.add_shift_type('Shift B', '08:00', '16:00')
    shift_id = sh.list_shift_types()[0]['id']
    ok, msg = sc.assign_shift(emp_id, shift_id, '2026-09-05', status='Worked', notes='')
    assert ok, msg

    client = app_module.app.test_client()
    response = client.get('/reports/export?year=2026&month=9')
    assert response.status_code == 200

    workbook = load_workbook(BytesIO(response.data))
    rows = list(workbook.active.iter_rows(values_only=True))
    header = rows[3]
    assert 'Monthly Hours' in header or 'Monthly Hours' in [h for h in header if h]
    assert 'Annual Hours' in header or 'Annual Hours' in [h for h in header if h]
    assert any(row and row[0] == 'E-300' and row[7] is not None for row in rows)


def test_reports_page_shows_total_annual_hours_for_all_employees():
    client = app_module.app.test_client()
    response = client.get('/reports?year=2026&month=9')
    assert response.status_code == 200
    html = response.get_data(as_text=True)
    assert 'Total Annual Hours' in html
    assert 'Total Monthly Hours' in html


def test_reports_export_includes_total_monthly_and_annual_rows(tmp_path, monkeypatch):
    setup_test_db(tmp_path, monkeypatch)

    ok, msg = emp.add_employee('E-400', 'Delta', 'Four', department='Ops', position='Nurse')
    assert ok, msg
    ok, msg = emp.add_employee('E-401', 'Echo', 'Five', department='Ops', position='Nurse')
    assert ok, msg
    sh.add_shift_type('Shift C', '08:00', '16:00')
    shift_id = sh.list_shift_types()[0]['id']
    for employee in emp.list_employees(status='Active'):
        ok, msg = sc.assign_shift(employee['id'], shift_id, '2026-09-05', status='Worked', notes='')
        assert ok, msg

    client = app_module.app.test_client()
    response = client.get('/reports/export?year=2026&month=9')
    assert response.status_code == 200

    workbook = load_workbook(BytesIO(response.data))
    rows = list(workbook.active.iter_rows(values_only=True))
    assert any(row and row[0] == 'Total Monthly Hours' for row in rows)
    assert any(row and row[0] == 'Total Annual Hours' for row in rows)


