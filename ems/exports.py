"""
exports.py
Export employee hours summaries to an Excel (.xlsx) workbook — one row
per active employee for a given month, covering the same figures shown
on the Reports page (worked hours, leave by category, OT breakdown,
total paid hours). Used by all three front ends so "download the month
as a spreadsheet" works the same way everywhere.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from ems import employees as emp_mod
from ems import reports as rp

HEADER_FILL = PatternFill(start_color="1B2430", end_color="1B2430", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(size=10, italic=True, color="4A5568")
THIN_BORDER = Border(bottom=Side(style="thin", color="DEDBD3"))

COLUMNS = [
    ("employee_code", "Code", 10),
    ("name", "Employee", 22),
    ("department", "Department", 16),
    ("position", "Position", 16),
    ("eft", "EFT", 8),
    ("expected_hours_monthly", "Monthly Hours", 12),
    ("expected_hours_annually", "Annual Hours", 12),
    ("worked_hours", "Worked Hrs", 12),
    ("sick_hours", "Sick Hrs", 10),
    ("personal_hours", "Personal Hrs", 12),
    ("vacation_hours", "Vacation Hrs", 12),
    ("other_leave_hours", "Other Leave Hrs", 14),
    ("unpaid_leave_hours", "Unpaid Leave Hrs", 15),
    ("total_leave_hours", "Total Leave Hrs", 14),
    ("manual_overtime_hours", "Manual OT Hrs", 13),
    ("auto_overtime_daily", "Auto OT (Daily)", 14),
    ("auto_overtime_biweekly", "Auto OT (Biweekly)", 17),
    ("overtime_hours", "Total OT Hrs", 12),
    ("total_paid_hours", "Total Paid Hrs", 14),
]


def _row_for_employee(employee_id, year, month):
    """Flatten one employee_hours_breakdown() result into the flat set of
    values COLUMNS expects, pulling employee identity fields in too."""
    data = rp.employee_hours_breakdown(employee_id, year, month)
    if not data:
        return None
    e = data["employee"]
    row = dict(data)
    row["employee_code"] = e["employee_code"]
    row["name"] = f"{e['first_name']} {e['last_name']}"
    row["department"] = e["department"] or ""
    row["position"] = e["position"] or ""
    row["eft"] = e["eft"]
    row["expected_hours_monthly"] = data.get("expected_hours_monthly")
    row["expected_hours_annually"] = data.get("expected_hours_annually")
    row["worked_hours"] = data.get("worked_hours")
    row["sick_hours"] = data.get("sick_hours")
    row["personal_hours"] = data.get("personal_hours")
    row["vacation_hours"] = data.get("vacation_hours")
    row["other_leave_hours"] = data.get("other_leave_hours")
    row["unpaid_leave_hours"] = data.get("unpaid_leave_hours")
    row["total_leave_hours"] = data.get("total_leave_hours")
    row["manual_overtime_hours"] = data.get("manual_overtime_hours")
    row["auto_overtime_daily"] = data.get("auto_overtime_daily")
    row["auto_overtime_biweekly"] = data.get("auto_overtime_biweekly")
    row["overtime_hours"] = data.get("overtime_hours")
    row["total_paid_hours"] = data.get("total_paid_hours")
    return row


def export_monthly_summary_to_excel(year, month, out_path, status="Active", employee_id=None):
    """Write one workbook to out_path with a summary row per employee
    (default: active employees only) for the given month. Returns
    out_path. Column set matches the Reports page plus a few extra
    identity/leave-category columns for a fuller offline record. If
    employee_id is supplied, export only that employee to match the
    filtered report view."""
    employees = emp_mod.list_employees(status=status) if status else emp_mod.list_employees()
    if employee_id is not None:
        employees = [e for e in employees if e["id"] == employee_id]

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    month_label = rp.month_name(year, month)
    ws.merge_cells(f"A1:{get_column_letter(len(COLUMNS))}1")
    ws["A1"] = f"Employee Hours Summary — {month_label}"
    ws["A1"].font = TITLE_FONT

    ws.merge_cells(f"A2:{get_column_letter(len(COLUMNS))}2")
    ws["A2"] = f"Generated from Staffing Board · {len(employees)} employee(s) · status filter: {status or 'All'}"
    ws["A2"].font = SUBTITLE_FONT

    header_row = 4
    for col_idx, (key, label, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    r = header_row + 1
    total_annual_hours = 0.0
    total_monthly_hours = 0.0
    for e in employees:
        row = _row_for_employee(e["id"], year, month)
        if not row:
            continue
        total_annual_hours += float(row.get("expected_hours_annually") or 0)
        total_monthly_hours += float(row.get("expected_hours_monthly") or 0)
        for col_idx, (key, label, width) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=r, column=col_idx, value=row.get(key))
            cell.border = THIN_BORDER
            if key not in ("employee_code", "name", "department", "position"):
                cell.alignment = Alignment(horizontal="right")
        r += 1

    last_row = r - 1
    if employees:
        annual_total_row = r
        monthly_total_row = r + 1
        ws.cell(row=annual_total_row, column=1, value="Total Annual Hours")
        ws.cell(row=annual_total_row, column=7, value=round(total_annual_hours, 2))
        ws.cell(row=monthly_total_row, column=1, value="Total Monthly Hours")
        ws.cell(row=monthly_total_row, column=6, value=round(total_monthly_hours, 2))

        for total_row in (annual_total_row, monthly_total_row):
            for col_idx in range(1, len(COLUMNS) + 1):
                cell = ws.cell(row=total_row, column=col_idx)
                if col_idx in (1, 6, 7):
                    cell.font = HEADER_FONT
                    cell.fill = HEADER_FILL
                cell.border = THIN_BORDER
        last_row = monthly_total_row
        r = monthly_total_row + 1

    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(COLUMNS))}{last_row}"

    wb.save(out_path)
    return out_path
